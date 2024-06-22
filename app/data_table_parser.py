import re
from decimal import Decimal
import dateutil.parser
from app.account_entry import AccountEntry
import logging

import csv
import openpyxl
import xlrd
import json
from html_table_parser import HTMLTableParser

class DataTableParser:
    def __init__(self, filename):
        self.filename = filename
        self.header_row = None
        self.headers = {}
    
    def _get_rows(self):
        """
        Returns all rows in the file, whether they belong to the data table or not
        """
        raise NotImplementedError

    def _get_data_rows(self):
        """
        Returns the rows of the data table
        Does not return the header rows or anything preceding it
        """
        return list(self._get_rows())[self.header_row+1:]

    def get_headers(self):
        """
        Returns a dict of data table headers, with the keys being the header text, and the values being the column index
        """
        # First, find the header row
        for rowidx, row in enumerate(self._get_rows()):
            if len(self.headers) == 0:
                hcount = 0
                for colidx, cell in enumerate(row):
                    for field in AccountEntry.fields:
                        if cell in AccountEntry.fields[field]:
                            logging.debug(f'Found match for {field} in cell {rowidx}:{colidx} - {cell}')
                            self.headers[field] = colidx
                            hcount += 1
                if (hcount >= 3):
                    self.header_row = rowidx
                    logging.debug(f'Found header row at row {self.header_row}')
                    logging.debug(self.headers)
                    break
                else:
                    self.headers = {}
            
        return self.headers

    def get_entries(self):
        """
        Returns a list of AccountEntry items
        """
        if len(self.headers) == 0:
            self.get_headers()

        items = []
        for row in self._get_data_rows():
            rowData = {}
            for field in self.headers:
                if len(row) > self.headers[field]:
                    rowData[field] = row[self.headers[field]]
            
            curItem = self.parseItem(rowData)
            if curItem is not None:
                logging.debug(curItem)
                items.append(curItem)

        return items

    def parseNumber(self, amount):
        if amount is None or amount == '--' or amount == '0' or amount == '0.0' or amount == '':
            return Decimal(0)

        if type(amount) == str:
            amount = re.sub(r'[^\d\.]+', '', amount)
            return Decimal(amount)
        elif type(amount) == float:
            return Decimal(str(amount))

        return amount

    def parseItem(self, item):
        curItem = AccountEntry()

        for field in AccountEntry.fields:
            if field in item and item[field] is not None and item[field] != '':
                if type(item[field]) == str:
                    item[field] = re.sub(r'^\((.*)\)$', r'-\1', item[field])
                    item[field] = item[field].replace(',', '')
                elif type(item[field]) == float:
                    item[field] = Decimal(str(item[field]))
                curItem.__setattr__(field, item[field])

        if 'date' not in item or item['date'] is None or item['date'] == '' or item['date'] == 'Pending' or item['date'] == '--':
            return None

        if type(item['date']) == str:
            try:
                curItem.date = dateutil.parser.isoparse(item['date'])
            except:
                curItem.date = dateutil.parser.parse(item['date'], dayfirst=True)

        if 'withdrawal' in item and 'deposit' in item:
            curItem.amount = self.parseNumber(item['deposit']) - self.parseNumber(item['withdrawal'])
        elif 'withdrawal' in item:
            amount = self.parseNumber(item['withdrawal'])
            if amount == 0:
                return None

            if type in item and item['type'] == 'Debit':
                curItem.amount = -1 * amount
            else:
                curItem.amount = amount
        else:
            amount = self.parseNumber(item['amount'])
            if amount == 0:
                return None
            curItem.amount = amount

        return curItem

class XLSParser(DataTableParser):
    def __init__(self, filename):
        super().__init__(filename)
        self.wb = xlrd.open_workbook(self.filename)
        self.sh = self.wb.sheet_by_index(0)

    def _get_rows(self):
        return [[cell.value for cell in row] for row in list(self.sh.get_rows())]

class XLSXParser(DataTableParser):
    def __init__(self, filename):
        super().__init__(filename)
        self.wb = openpyxl.load_workbook(self.filename)
        self.sh = self.wb.active

    def _get_rows(self):
        return self.sh.iter_rows

    def _get_data_rows(self):
        yield self.sh.iter_rows(min_row=self.header_row+1)

class CSVParser(DataTableParser):
    def __init__(self, filename):
        super().__init__(filename)
        with open(self.filename) as infile:
            self.csvrows = list(csv.reader(infile, delimiter=','))

    def _get_rows(self):
        return [[x.strip() for x in row] for row in self.csvrows]

class HTMLParser(DataTableParser):
    def __init__(self, filename):
        super().__init__(filename)
        with open(filename, "r") as fh:
            xhtml = fh.read()
            self.parser = HTMLTableParser()
            self.parser.feed(xhtml)

        self.table = self.parser.tables[0]
    
    def get_headers(self):
        for self.table in self.parser.tables:
            super().get_headers()
            if len(self.headers) > 0:
                break

    def _get_rows(self):
        return self.table
    
class JSONParser(DataTableParser):
    def __init__(self, filename):
        super().__init__(filename)
        with open(self.filename) as infile:
            self.json_items = json.load(infile)
            if type(self.json_items) == dict and 'dtTrxnResult' in self.json_items:
                self.json_items = self.json_items['dtTrxnResult']

    def _get_rows(self):
        return self.json_items

